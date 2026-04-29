import requests
import pandas as pd
import time
import os
import json
from datetime import datetime

# =========================
# CONFIG
# =========================
BASE_URL = "https://insp.social/xrpc/app.bsky.feed.getFeed"
FEED = "at://did:plc:cadeggccwqtp3yjkk7auhpil/app.bsky.feed.generator/innovator"
LIMIT = 50

raw_token = os.getenv("INSPIRED_TOKEN", "")
token = raw_token.replace("\n", "").replace("\r", "").strip()

if not token:
    print("❌ TOKEN NOT FOUND / EMPTY")
    exit()

print("✅ Token loaded. Length:", len(token))

headers = {
    "Authorization": f"Bearer {token}",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept": "application/json",
    "Connection": "keep-alive"
}

timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
file_path = f"inspired_data_{timestamp}.xlsx"
cache_file = "last_success.json"

# =========================
# FETCH DATA (WITH BETTER RETRY)
# =========================
all_rows = []
cursor = None
MAX_RETRIES = 3

while True:
    params = {
        "feed": FEED,
        "limit": LIMIT
    }
    if cursor:
        params["cursor"] = cursor

    success = False

    for attempt in range(MAX_RETRIES):
        try:
            res = requests.get(BASE_URL, headers=headers, params=params, timeout=10)

            if res.status_code == 200:
                success = True
                break
            else:
                print(f"⚠️ Attempt {attempt+1}: {res.status_code}")
                print(res.text[:200])

        except Exception as e:
            print(f"⚠️ Attempt {attempt+1} error:", str(e))

        time.sleep(2 * (attempt + 1))

    if not success:
        print("❌ Failed after retries. Stop fetching.")
        break

    data = res.json()
    feed = data.get("feed", [])

    if not feed:
        break

    for item in feed:
        post = item.get("post", {})
        author = post.get("author", {})

        handle = author.get("handle", "")
        safe_handle = handle.split('.')[0] if handle and '.' in handle else handle

        all_rows.append({
            "Name": author.get("displayName"),
            "Handle": handle,
            "Profile": f"https://app.inspired.ch/profile/{safe_handle}" if handle else "",
            "Country": author.get("country"),
            "Sector ID": author.get("sectorId"),
            "Stage ID": author.get("stageId"),
            "User Type": author.get("userType"),
            "Description": post.get("record", {}).get("text") or "",
            "Likes": post.get("likeCount") or 0,
        })

    cursor = data.get("cursor")
    if not cursor:
        break

    time.sleep(1)

print(f"✅ Total data fetched: {len(all_rows)}")

# =========================
# FALLBACK STRATEGY (SMART)
# =========================
if len(all_rows) == 0:
    print("⚠️ API failed. Trying fallback from last_success.json")

    if os.path.exists(cache_file):
        with open(cache_file, "r") as f:
            all_rows = json.load(f)
        print(f"✅ Loaded fallback data: {len(all_rows)} rows")
    else:
        print("❌ No fallback available. Using minimal placeholder")
        all_rows.append({
            "Name": "NO DATA",
            "Handle": "-",
            "Profile": "-",
            "Country": "-",
            "Sector ID": "ERROR",
            "Stage ID": "-",
            "User Type": "-",
            "Description": "API failed and no cache available",
            "Likes": 0,
        })
else:
    # SAVE CACHE
    with open(cache_file, "w") as f:
        json.dump(all_rows, f)
    print("💾 Saved last successful fetch")

# =========================
# SAVE TO EXCEL
# =========================
df = pd.DataFrame(all_rows)
df.to_excel(file_path, index=False)

# =========================
# STYLING
# =========================
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook(file_path)
ws = wb.active

# HEADER
header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# BODY
body_font = Font(name="Segoe UI", size=10)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# WIDTH
column_widths = {
    "Name": 28,
    "Description": 60,
    "Profile": 35,
    "Handle": 18,
    "Country": 12,
    "Sector ID": 14,
    "Stage ID": 14,
    "User Type": 14,
    "Likes": 10,
}

for idx, cell in enumerate(ws[1], start=1):
    ws.column_dimensions[get_column_letter(idx)].width = column_widths.get(cell.value, 15)

# HEIGHT
for row in ws.iter_rows(min_row=2):
    ws.row_dimensions[row[0].row].height = 60

ws.freeze_panes = "A2"

# ZEBRA
stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    if i % 2 == 0:
        for cell in row:
            cell.fill = stripe_fill

# SAVE FINAL
wb.save(file_path)

print(f"🔥 DONE — FILE CREATED: {file_path}")
